#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''=================================================
@Project -> File   ：test -> get_info_from_excel
@IDE    ：PyCharm
@Author ：zhang yigang

@Date   ：19-7-25 下午3:02
@Desc   ：
=================================================='''
import openpyxl
import sys

reload(sys)
sys.setdefaultencoding("utf-8")
class myexcel():
    def __init__(self):
        self.excel_array = []
        self.excel_keys =[]
        self.merge_dict = {}  # 用于保存计算出的合并的单元格，key=(7, 4)合并单元格坐标，value=(7, 2)合并单元格首格坐标
        self.sheet = None

    def get_row_and_color(self,bounds):
        start_col = bounds[0]
        start_row = bounds[1]
        end_col = bounds[2]
        end_row = bounds[3]
        return start_col, start_row, end_col, end_row

    def get_merge(self, table):

        # 返回数据合并的列数与行数，四元组
        if table.merged_cells:
            # 一个元组，前两个是开始列行，后两个是结束列行
            for item in table.merged_cells:
                start_col, start_row, end_col, end_row = self.get_row_and_color(item.bounds)
                if start_col == end_col:
                    #     列相同,说明是行合并
                    for row in range(start_row + 1, end_row + 1):
                        # 取数是保存左闭右开，因此为了取后面的采取以上策略
                        same_key = str(start_col) + "_" + str(row)
                        self.merge_dict[same_key] = str(start_col) + "_" + str(start_row)
                elif start_row == end_row:
                    for col in range(start_col + 1, end_col + 1):
                        # 取数是保存左闭右开，因此为了取后面的采取以上策略
                        same_key = str(col) + "_" + str(start_row)
                        self.merge_dict[same_key] = str(start_col) + "_" + str(start_row)
                else:
                    for row in range(start_row + 1, end_row + 1):
                        for col in range(start_col + 1, end_col + 1):
                            same_key = str(col) + "_" + str(row)
                            self.merge_dict[same_key] = str(start_col) + "_" + str(start_row)

    def load_excel(self, filepath):
        # 打开excel文件,获取工作簿对象
        wb = openpyxl.load_workbook(filepath)

        # 从工作薄中获取一个表单(sheet)对象

        self.sheet = wb.get_active_sheet()
        self.get_merge(self.sheet)
        self.max_row = self.sheet.max_row
        self.max_col = self.sheet.max_column
        self.set_headers()
        for row in range(2,self.max_row+1):
            every_dict = {}
            # 从开始第二行,第1行作为header
            for col in range(1, self.max_col+1):
                 every_dict[self.excel_keys[col]] = self.get_value_from_row_and_col(row, col)
            self.excel_array.append(every_dict)

    def get_value_from_row_and_col(self, row, col):
        cell_value = self.sheet.cell(row,col).value
        if not cell_value:
            key = str(col) + "_" + str(row)
            if key in self.merge_dict:
                dict_value = self.merge_dict.get(key)
                col,row = dict_value.split("_")
                cell_value = self.sheet.cell(int(row),int(col)).value
            else:
                cell_value = ""
        cell_value = self.del_unuse_char(cell_value)
        return cell_value

    def del_unuse_char(self,single_info):
        import re
        single_info = str(single_info)
        single_info = single_info.replace("\n", "")
        if single_info.strip() == "":
            return ""
        patten = re.compile(r"\s{2,}")
        string = patten.sub(" ", single_info)
        return string.strip()

    def set_headers(self):
        self.excel_keys.append(" ")
        # 加这个是因为excle取是从1开始的，为了好对应
        for col in range(1, self.max_col+1):
            self.excel_keys.append(self.sheet.cell(1,col).value)

    def get_excel_json(self):
        import json
        if len(self.excel_array)!=0:
            return json.dumps(self.excel_array)

    def get_all_value_key(self, key):
        value_array = []
        if not key in self.excel_keys:
            return False
        for every_dict in self.excel_array:
            value_array.append(every_dict[key])
        return value_array

if __name__=="__main__":
    myexcel = myexcel()
    myexcel.load_excel("/Users/bupt/Documents/alading.xlsx")
    print myexcel.get_excel_json()
